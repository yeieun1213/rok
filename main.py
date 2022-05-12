import pyautogui
import pytesseract
import clipboard
from openpyxl import Workbook, load_workbook
import time
from PIL import Image
#import cv2
import os

today = time.strftime('%m%d', time.localtime(time.time()))

# 파일 경로
root = 'C:/Users/yeieu/Desktop/rok/'
pytesseract.pytesseract.tesseract_cmd = R'C:/Program Files/Tesseract-OCR/tesseract'

# 시작, 
#rank = int(input("start rank:"))
#goal = int(input("goal rank:"))
rank, goal = 431, 435
x, y = 160, 320

# to print a current cursor position
def position():
    while(True):
        print(pyautogui.position())
        time.sleep(1)

# SETTING of start position
def setting():
    x, y = pyautogui.position()

# to make excel file and set a base row
def makexl(list):
    try:
        # 파일 경로
        write_wb = load_workbook('C:/Users/yeieu/Desktop/rok/0911.xlsx')
    except:
        write_wb = Workbook()   # make excel file
    write_ws = write_wb.create_sheet(today) # make new sheet with name

    write_ws = write_wb.active
    for i in range(len(list)):
        row = chr(66+i)+'1'     # unicode of 'B' = 66
        write_ws[row] = list[i] # create row
    return write_wb, write_ws

# to append result into excel file
def addcolumn(ws, list):
    for i in range(len(list)):
        ws.cell(row=rank+1, column=1).value = rank
        ws.cell(row=rank+1, column=i+2).value = list[i]
    return ws

# to find a user in auto-scrolling with 'pyautogui'.
# it makes the point move, drag and click automatically.
def find_user():
    if rank > 3:
        y = 320
        pyautogui.moveTo(x, y)  # move to 'rank'th ranking position
        time.sleep(0.3)
        pyautogui.dragTo(x, y - 60, 2, pyautogui.easeOutQuad)  # scroll for going to 'rank+1'th rank
        pyautogui.moveTo(x, y)  # move to 'rank+1'th rank
        time.sleep(1)
        pyautogui.click(x, y)  # click for information of users
        time.sleep(2)
        return
    elif rank == 1:
        y = 210
        pyautogui.moveTo(x, y)
        time.sleep(1)
        pyautogui.click(x, y)  # click for information of users
        time.sleep(2)
        return
    elif rank == 2:
        y = 260
        pyautogui.moveTo(x, y)
        time.sleep(1)
        pyautogui.click(x, y)  # click for information of users
        time.sleep(2)
        return
    elif rank == 3:
        y = 310
        pyautogui.moveTo(x, y)
        time.sleep(1)
        pyautogui.click(x, y)  # click for information of users
        time.sleep(2)
        return

def capture(option):

    if option == "uid":
        uidImg = root + str(rank) + '_uid.png'
        pyautogui.screenshot(uidImg, region=(419, 173, 80, 20))
        return uidImg
    elif option == "death":
        deathImg = root + str(rank) + '_death.png'
        pyautogui.screenshot(deathImg, region=(661, 289, 70, 50))
        return deathImg

    elif option == "power":
        powerImg = root + str(rank) + '_power.png'
        pyautogui.screenshot(powerImg, region=(503, 246, 100, 30))
        return powerImg
    elif option == "point":
        pointImg = root + str(rank) + '_point.png'
        pyautogui.screenshot(pointImg, region=(628, 245, 120, 30))
        return pointImg
    elif option == "detailT4":
        detailT4Img = root + str(rank) + '_detailT4.png'
        pyautogui.screenshot(detailT4Img, region=(487, 430, 60, 20))
        return detailT4Img
    elif option == "detailT5":
        detailT5Img = root + str(rank) + '_detailT5.png'
        pyautogui.screenshot(detailT5Img, region=(487, 454, 60, 20))
        return detailT5Img

# OCR of tesseract
def OCR(img,option=""):
    # image preprocessiong: change the color photo to the gray-scale photo
    imgGray = Image.open(img).convert('L')
    return pytesseract.image_to_string(imgGray, config='--psm 1 -c preserve_interword_spaces=1')
    #imgGray = cv2.cvtColor(cv2.imread(img), cv2.COLOR_RGB2GRAY)

    #clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
    #gray = clahe.apply(imgGray)

    #Img = cv2.imread(img)
    #imggray = cv2.cvtColor(Img, cv2.COLOR_RGB2GRAY)
    #noise = cv2.threshold(imggray, 0, 200, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]
    #gray = noise


    #return pytesseract.image_to_string(gray)
    #return pytesseract.image_to_string(gray, config='--psm 1 -c preserve_interword_spaces=1')


list = []
def test():
    # start
    list.clear()
    print("rank:", rank)

    # screenshot
    find_user()
    time.sleep(1)
    
    # uid
    img = capture('uid')
    uid = OCR(img).split(')')[0]
    print('uid:', uid)
    list.append(uid)
    os.remove(img)
    time.sleep(0.3)
    
    # nickname
    pyautogui.click(374, 205) # click for nickname
    time.sleep(0.3)
    name = clipboard.paste()
    print('name:', name)
    list.append(name)
    time.sleep(0.3)
    
    # power
    img = capture('power')
    time.sleep(0.3)
    power = OCR(img).split('\n')[0]
    if power[-1] == ',':
        power = power[:-2]
    print('power:', power)
    list.append(power)
    os.remove(img)
    time.sleep(0.5)

    # kill points
    img = capture('point')
    time.sleep(0.3)
    point = OCR(img, 'point').split('\n')[0]
    if point[-1] == ',':
        point = point[:-2]
    print('point:', point)
    list.append(point)
    os.remove(img)
    time.sleep(0.5)
    
    # detail of kill points
    pyautogui.click(627, 241)  # click for detail
    time.sleep(0.5)

    # screenshot
    path = root + str(rank) + '.png'
    pyautogui.screenshot(path, region=(115,86,700, 420))
    time.sleep(0.7)

    # kills for T4
    img = capture('detailT4')
    time.sleep(0.5)
    detailT4 = OCR(img, 'detailT4').split('\n')[0]
    if detailT4[-1] == ',':
        detailT4 = detailT4[:-2]
    print('T4:', detailT4)
    list.append(detailT4)
    os.remove(img)
    time.sleep(0.5)

    # kills for T5
    img = capture('detailT5')
    time.sleep(0.3)
    detailT5 = OCR(img, 'detailT5').split('\n')[0]
    if detailT5[-1] == ',':
        detailT5 = detailT5[:-2]
    print('T5:', detailT5)
    list.append(detailT5)
    os.remove(img)
    time.sleep(0.5)


    # death
    pyautogui.click(223, 409) # click for detail of death
    time.sleep(0.7)
    img = capture('death')
    time.sleep(0.3)
    death = OCR(img).split('\n')[0]
    if death[-1] == ',':
        death = death[:-2]
    print('death:', death)
    list.append(death)
    os.remove(img)

    # screenshot
    path = root + str(rank) + '_death.png'
    pyautogui.screenshot(path, region=(95, 60, 700, 480))
    time.sleep(0.7)
    
    pyautogui.click(870, 70)
    time.sleep(0.5)

    return list

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    position()
    setting()

    wb, ws = makexl(['uid', 'name', 'power', 'point', 'T4', 'T5', 'death'])

    for i in range(rank, goal+1):
        try:
            list = test()
            print(list)
            pyautogui.click(762, 104) # end
            time.sleep(1)
            try:
                ws = addcolumn(ws, list)
            except:
                l = []
                for i in range(len(list)):
                    l.append(0)
                ws = addcolumn(ws, l)
            rank += 1
        
        # 오류 생기면 그냥 저장
        finally:
            path = 'C:/Users/yeieu/Desktop/rok/0911.xlsx'
            wb.save(path)

    print("-------------end------------")


    #position()
